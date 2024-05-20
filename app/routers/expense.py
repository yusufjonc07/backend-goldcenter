from click import Option
from fastapi import Body, HTTPException, APIRouter, Depends, UploadFile, File, Form
from typing import Optional
import json

from pydantic import ValidationError
from app.models.contragent import Contragent
from app.schemas.user import NewUser
from app.utils.fileUtil import save_file, validate_file
from app.utils.handler import integrityHandler
from security.auth import get_current_active_user
from databases.main import ActiveSession
from sqlalchemy.orm import Session
from app.models.expense import *
from app.functions.expense import *
from app.schemas.expense import *
from datetime import date
from openpyxl import load_workbook
from pydantic import ValidationError

expense_router = APIRouter(tags=['Kassa Endpoint'])


@expense_router.get("/expenses", description="This router returns list of the expenses using pagination")
async def get_expenses_list(
    search: Optional[str] = "",
    expenseType: Optional[ExpenceTypes] = None,
    fromDate: Optional[date] = None,
    toDate: Optional[date] = None,
    employeeId: Optional[int] = 0,
    contragentId: Optional[int] = 0,
    moneyFormId: Optional[int] = 0,
    page: int = 1,
    limit: int = 10,
    db: Session = ActiveSession,
    usr: NewUser = Depends(get_current_active_user)
):
    if not usr.userRole in ['any_role']:
        return get_all_expenses(moneyFormId, search, expenseType, fromDate, toDate, employeeId, contragentId,  page, limit, usr, db)
    else:
        raise HTTPException(status_code=400, detail="Sizga ruxsat berilmagan!")


@expense_router.post("/expense/create")
async def create_new_income(
    form_data_string: str = Form(...,
                                 description="Bu yerga string kelishi kerak"),
    file: Optional[UploadFile] = File("none"),
    db: Session = ActiveSession,
    usr: User = Depends(get_current_active_user)
):
    """
        `form_data_string`
        [
            {
                "type": "salary",
                "employeeId": 12,
                "contragentId": 0,
                "value": 1,
                "moneyFormId": 1,
                "isAvanse": false,
                "comment": "string"
            }
        ]
    """

    if not usr.userRole in ['any_role']:

        try:
            form_datas_dict = json.loads(form_data_string)
        except Exception as err:
            raise HTTPException(
                422, f"form_data_string xato jo'natildi: {err.args}")

        if type(form_datas_dict).__name__ != 'list':
            raise HTTPException(
                422, "form_data_string - list emas, object kelyapti")

        form_datas = []
        for form_data_dict in form_datas_dict:
            try:
                form_datas.append(
                    NewExpense(**form_data_dict)
                )
            except ValidationError as vErr:
                raise HTTPException(422, vErr.errors())

        if file != 'none':
            fileName = await validate_file(file, ['document', 'image'], 3)
        else:
            fileName = None

        for form_data in form_datas:

            print(form_data)

            if form_data.type == 'salary':
                employee = db.get(Employee, form_data.employeeId)

                if not employee:
                    raise HTTPException(400, "Hodim topilmadi topilmadi")
                else:
                    employee.balance -= form_data.value
                    db.flush()

            if form_data.type == 'contragent':
                contragent = db.get(
                    Contragent, form_data.contragentId)

                if not contragent:
                    raise HTTPException(
                        400, f"Doimiy chiqim topilmadi")
                else:
                    contragent.balance -= form_data.value
                    db.flush()
            try:

                if not form_data.contragentId:
                    form_data.contragentId = 0

                new_expense = Expense(
                    type=form_data.type,
                    employeeId=form_data.employeeId if form_data.employeeId > 0 else None,
                    contragentId=form_data.contragentId if form_data.contragentId > 0 else None,
                    value=form_data.value,
                    moneyFormId=form_data.moneyFormId,
                    branchId=usr.branchId,
                    comment=form_data.comment,
                    userId=usr.id,
                    isAvanse=form_data.isAvanse,
                    fileName=fileName,
                )

                db.add(new_expense)
                db.flush()

                if new_expense.moneyForm.balance < new_expense.value:
                    raise HTTPException(
                        400, f"{new_expense.moneyForm.name} balansida yetarli mablag' yetarli emas!")
                else:
                    new_expense.moneyForm.balance -= new_expense.value

            except IntegrityError as e:
                raise integrityHandler(e)

        if fileName:
            await save_file(file, fileName, f"expenses")

        db.commit()

        raise HTTPException(200, "Ma`lumotlar saqlandi!")

    else:
        raise HTTPException(status_code=400, detail="Sizga ruxsat berilmagan!")


@expense_router.post("/expense/excelUpload")
async def upload_one_excel(
    file: UploadFile = File(...),
    db: Session = ActiveSession,
    usr: NewUser = Depends(get_current_active_user)
):
    """
    1000 qatorgacha excel fayl yuklash mumkin
    <a target="blank" href='https://gc-api.magnitgroup.uz/files/excel/HODIM_MAOSH_SHABLON.xlsx'>SHABLON👈</a>
    """

    if not usr.userRole in ['any_role']:

        # validate file
        await validate_file(file, ['document'], 3)

        # save file
        fileUploadUrl = await save_file(file, file.filename, f"")

        # load excel data as a workbook
        wb = load_workbook(filename=fileUploadUrl+file.filename)

        # get active sheet
        ws = wb.active

        # validate worksheet
        validate_excel_ws(ws, db)

        form_datas = []
        for row in ws.iter_rows(min_row=2, max_col=7, max_row=1000):

            # stop a loop when the cell is empty
            if row[0].value is None:
                break

            try:
                form_datas.append(
                    NewExpenseExcel(
                        pnfl=row[0].value,
                        value=row[1].value,
                        moneyFormName=row[2].value,
                        date=row[3].value,
                        isAvanse=row[4].value,
                    )
                )
            except ValidationError as vErr:
                raise HTTPException(422, vErr.errors())

        return create_expense_by_ecxel_data(form_datas, db, usr)
    else:
        raise HTTPException(status_code=400, detail="Sizga ruxsat berilmagan!")


@expense_router.put("/expense/{id}/update", description="This router is able to update expense")
async def update_one_expense(
    id: int,
    form_data: UpdateExpense,
    db: Session = ActiveSession,
    usr: NewUser = Depends(get_current_active_user)
):
    if not usr.userRole in ['any_role']:
        return update_expense(id, form_data, usr, db)
    else:
        raise HTTPException(status_code=400, detail="Sizga ruxsat berilmagan!")
