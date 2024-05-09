from fastapi import HTTPException, APIRouter, Depends, UploadFile, File
from typing import Optional, List
from app.schemas.user import NewUser
from security.auth import get_current_active_user
from databases.main import ActiveSession
from sqlalchemy.orm import joinedload, Session
from app.models.income import *
from app.functions.income import *
from app.schemas.income import *
from app.functions.moneyHistory import get_all_agreement_payments
from datetime import date
from app.utils.fileUtil import save_file, validate_file
from openpyxl import load_workbook
from pydantic import ValidationError


income_router = APIRouter(tags=['Kassa Endpoint'])


@income_router.get("/incomes")
async def get_agreement_payments(
    clientId: Optional[int] = 0,
    contragentId: Optional[int] = 0,
    floorId: Optional[int] = 0,
    moneyFormId: Optional[int] = 0,
    type: Optional[IncomeType] = None,
    fromDate: Optional[date] = None,
    toDate: Optional[date] = None,
    page: int = 1,
    limit: int = 10,
    db: Session = ActiveSession,
    usr: NewUser = Depends(get_current_active_user)
):
    if not usr.userRole in ['any_role']:
        return get_all_agreement_payments(moneyFormId, clientId, contragentId, floorId, type, fromDate, toDate, page, limit, usr, db)
    else:
        raise HTTPException(status_code=400, detail="Sizga ruxsat berilmagan!")


@income_router.post("/income/create", description="This router is able to add new income")
async def create_new_income(
    form_datas: List[NewIncome],
    db: Session = ActiveSession,
    usr: NewUser = Depends(get_current_active_user)
):
    if not usr.userRole in ['any_role']:
        return create_income(form_datas, usr, db)
    else:
        raise HTTPException(status_code=400, detail="Sizga ruxsat berilmagan!")


@income_router.put("/income/{id}/update", description="This router is able to update income")
async def update_one_income(
    id: int,
    form_data: UpdateIncome,
    db: Session = ActiveSession,
    usr: NewUser = Depends(get_current_active_user)
):
    if not usr.userRole in ['any_role']:
        return update_income(id, form_data, usr, db)
    else:
        raise HTTPException(status_code=400, detail="Sizga ruxsat berilmagan!")


@income_router.post("/income/excelUpload")
async def upload_one_excel(
    file: UploadFile = File(...),
    db: Session = ActiveSession,
    usr: NewUser = Depends(get_current_active_user)
):
    """
    1000 qatorgacha excel fayl yuklash mumkin
    <a target="blank" href='https://gc-api.magnitgroup.uz/files/excel/MIJOZ_KIRIM_SHABLON.xlsx'>SHABLON👈</a>
    """

    if not usr.userRole in ['any_role']:

        # validate file
        await validate_file(file, ['document'], 3)

        # save file
        fileUploadUrl = await save_file(file, file.filename, f"")

        # load excel data as a workbook
        wb = load_workbook(filename=fileUploadUrl+file.filename)

        # get active sheet
        ws = wb.active

        # validate worksheet
        validate_excel_ws(ws, db)

        form_datas = []
        for row in ws.iter_rows(min_row=2, max_col=7, max_row=1000):

            # stop a loop when the cell is empty
            if row[0].value is None:
                break

            try:
                form_datas.append(
                    NewIncomeExcel(
                        inn=row[0].value,
                        value=row[1].value,
                        moneyFormName=row[2].value,
                        date=row[3].value,
                        type=row[4].value,
                        forYear=row[5].value,
                        forMonth=row[6].value,
                    )
                )
            except ValidationError as vErr:
                raise HTTPException(422, vErr.errors())

        return create_income_by_ecxel_data(form_datas, db, usr)
    else:
        raise HTTPException(status_code=400, detail="Sizga ruxsat berilmagan!")


@income_router.put("/income/{id}/fileUpload")
async def update_one_income(
    id: int,
    file: UploadFile = File(...),
    db: Session = ActiveSession,
    usr: NewUser = Depends(get_current_active_user)
):

    if not usr.userRole in ['any_role']:
        fileName = await validate_file(file, ['document', 'image'], 3)
        income = db.query(Income).filter_by(id=id).first()
        income.fileName = fileName
        db.commit()
        await save_file(file, fileName, f"expenses")
        raise HTTPException(200, "Ma`lumotlar saqlandi!")
    else:
        raise HTTPException(status_code=400, detail="Sizga ruxsat berilmagan!")
